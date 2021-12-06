
import openpyxl
from common.func.operation import my_add
import pytest
import allure

'''

# 读取工作簿
book = openpyxl.load_workbook("../data/params.xlsx")
# 读取工作表
sheet = book.active
sheet1 = book.get_sheet_by_name("test_add")

# 读取单元格
single_cell_1 = sheet["A1"]
print(single_cell_1.value)

single_cell_2 = sheet.cell(column=3,row=3)
print(single_cell_2.value)

# 读取多个单元格
single_cells = sheet["A1:C3"]
print(type(single_cells))
print(single_cells)
'''

def open_excel_feil(path):
    book = openpyxl.load_workbook(path)
    return book
def open_excel_sheet(book,sheet_name=None):
    if sheet_name == None :
        sheet = book.active
    else:
        sheet = book.get_sheet_by_name(sheet_name)
    return sheet
def get_cells_value(sheet,begin,end):
    cells_value = sheet[f"{begin}:{end}"]
    return cells_value

def get_testcae_data(path,sheet_name,begin,end):
    req_list = []
    book = open_excel_feil(path)
    sheet = open_excel_sheet(book,sheet_name)
    cells = get_cells_value(sheet,begin,end)
    for row in cells:
        data =[]
        for cell in row:
            data.append(cell.value)
        req_list.append(data)

    return req_list

# if __name__ == '__main__':
#     list = get_testcae_data("../data/params.xlsx","test_add","A1","C4")
#     print(list)


# path = r"../data/params.xlsx"

class TestExcelDrivenDemo():
    @pytest.mark.parametrize("x,y,req", get_testcae_data("../../data/params.xlsx", "test_add", "A1", "C4"))
    def test_my_add_by_excel(self,x,y,req):
        assert my_add(x,y) == req

    @allure.title("allure 测试用例名称 ")
    def test_allure_demo_1(self):
        print("test_allure_demo_1")



